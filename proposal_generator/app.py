"""
app.py - Streamlit UI for PPA/EPC Proposal Generator

Run:
    streamlit run proposal_generator/app.py --server.port 8510
"""

from __future__ import annotations

import json
import math
import os
import re as _re
import subprocess
import sys
import tempfile
from pathlib import Path

# Ensure parent dir is on sys.path so `from proposal_generator...` works from any CWD
_project_root = str(Path(__file__).resolve().parent.parent)
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

import streamlit as st
import yaml
from streamlit_sortables import sort_items

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------

LOGO_PATH = Path(__file__).parent / "logo.png"
st.set_page_config(
    page_title="提案資料ジェネレーター",
    page_icon="☀️",
    layout="wide",
)

BASE_DIR = Path(__file__).parent
PROFILES_PATH = BASE_DIR / "composition_profiles.yaml"

_SORTABLE_STYLE = (
    ".sortable-item { text-align: left !important; "
    "max-width: 55% !important; padding: 5px 12px !important; "
    "font-size: 0.85rem !important; }"
)

# Hide anchor link icons on headings and fullscreen buttons on images
st.markdown("""
<style>
a.headerlink, .stMainBlockContainer [data-testid="StyledFullScreenButton"],
h1 a, h2 a, h3 a, .stMarkdown a[href^="#"] {
    display: none !important;
}
</style>
""", unsafe_allow_html=True)
EXCEL_PATH = BASE_DIR.parent / "ＰＬ_補ありなしPPAEPC_260317_XXXX様_v3.3.1.xlsm"
SAVE_DIR = BASE_DIR / "saved_cases"
SAVE_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Load profiles
# ---------------------------------------------------------------------------

def load_profiles():
    with open(PROFILES_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)

profiles_data = load_profiles()
profiles = profiles_data["profiles"]
catalog = profiles_data["slide_catalog"]

# ---------------------------------------------------------------------------
# Salesforce helpers
# ---------------------------------------------------------------------------

_SF_CMD = os.path.join(
    os.path.expanduser("~"), "AppData", "Roaming", "npm", "sf.cmd"
)
if not Path(_SF_CMD).exists():
    _SF_CMD = "sf"


def _sf_query(query: str, timeout: int = 20) -> list[dict]:
    """Run a SOQL query via sf CLI and return records list."""
    # Windows .cmd files require shell=True.
    # Escape % as %% to prevent cmd.exe env-var expansion in LIKE clauses.
    escaped_query = query.replace("%", "%%")
    cmd = f'"{_SF_CMD}" data query --query "{escaped_query}" --json'
    # CREATE_NO_WINDOW prevents DLL init failure (0xC0000142) when called
    # from Streamlit's process context on Windows.
    kwargs: dict = dict(capture_output=True, timeout=timeout, shell=True)
    if os.name == "nt":
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    result = subprocess.run(cmd, **kwargs)
    stdout_bytes = result.stdout
    if not stdout_bytes:
        stderr = result.stderr.decode("utf-8", errors="replace") if result.stderr else ""
        raise RuntimeError(f"sf query returned no stdout (rc={result.returncode}): {stderr[:200]}")
    out = stdout_bytes.decode("utf-8", errors="replace")
    data = json.loads(out)
    return data.get("result", {}).get("records", [])


def _tokenize_keyword(kw: str) -> list[str]:
    """Split keyword into tokens by spaces, parentheses, slashes etc."""
    normalized = _re.sub(r'[（）\(\)\s　/／・_＿]+', ' ', kw)
    tokens = [t.strip() for t in normalized.split() if t.strip()]
    return tokens if tokens else [kw]


@st.cache_data(ttl=None)
def _parse_quote_excel(uploaded_file) -> dict:
    """Parse a quote Excel file (当社標準テンプレート) and extract equipment + pricing.

    Reads from:
      - 入力② sheet: panels (row 11-12), PCS (row 13-16), battery (row 89)
      - ①表紙 sheet: pricing (AQ26-AQ31, AR28-AR29)
    """
    import openpyxl
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    result = {}

    # ---- 入力② sheet ----
    ws2 = None
    for name in wb.sheetnames:
        if "入力" in name and "②" in name:
            ws2 = wb[name]
            break
    if ws2:
        # System capacity
        result["pv_kw"] = ws2["C1"].value or 0
        result["pcs_kw"] = ws2["C2"].value or 0

        # Panels (rows 11-12)
        panels = []
        for r in [11, 12]:
            count = ws2.cell(r, 10).value  # J = count
            if count and int(count) > 0:
                panels.append({
                    "text": ws2.cell(r, 3).value or "",  # C = full text
                    "model": ws2.cell(r, 5).value or "",  # E = model
                    "maker": (ws2.cell(r, 6).value or "").replace("_", " "),  # F = maker
                    "output": ws2.cell(r, 7).value or "",  # G = output (e.g. "460W")
                    "unit_cost": ws2.cell(r, 9).value or 0,  # I = cost per unit
                    "count": int(count),
                })
        result["panels"] = panels
        result["panel_count"] = sum(p["count"] for p in panels)

        # PCS (rows 13-16)
        pcs_list = []
        for r in range(13, 17):
            count = ws2.cell(r, 10).value  # J = count
            if count and int(count) > 0:
                pcs_list.append({
                    "text": ws2.cell(r, 3).value or "",
                    "model": ws2.cell(r, 5).value or "",
                    "maker": (ws2.cell(r, 6).value or "").replace("_", " "),
                    "output": ws2.cell(r, 7).value or "",
                    "unit_cost": ws2.cell(r, 9).value or 0,
                    "count": int(count),
                })
        result["pcs_list"] = pcs_list
        result["pcs_count"] = sum(p["count"] for p in pcs_list)

        # Battery (row 89 = industrial, row 86 = residential)
        # Only include if J column (quantity) has a value > 0
        batteries = []
        for r in [89, 86]:
            qty = ws2.cell(r, 10).value  # J = quantity
            try:
                qty_i = int(qty) if qty is not None else 0
            except (ValueError, TypeError):
                qty_i = 0
            if qty_i <= 0:
                continue
            cap = ws2.cell(r, 7).value  # G = capacity (kWh)
            try:
                cap_f = float(cap) if cap is not None else 0
            except (ValueError, TypeError):
                cap_f = 0
            if cap_f > 0:
                batteries.append({
                    "text": ws2.cell(r, 3).value or "",
                    "model": ws2.cell(r, 5).value or "",
                    "maker": (ws2.cell(r, 6).value or "").replace("_", " "),
                    "kwh": cap_f,
                    "price": ws2.cell(r, 9).value or 0,  # I = price
                })
        result["batteries"] = batteries

    # ---- ①表紙 sheet ----
    ws1 = None
    for name in wb.sheetnames:
        if "表紙" in name and "①" in name:
            ws1 = wb[name]
            break
    if ws1:
        result["selling_price"] = ws1["AQ28"].value or 0
        result["kw_selling_price"] = ws1["AR28"].value or 0
        result["raw_cost"] = ws1["AQ29"].value or 0
        result["kw_unit_cost"] = ws1["AR29"].value or 0
        result["gross_profit"] = ws1["AQ30"].value or 0
        result["gross_margin_pct"] = (ws1["AQ31"].value or 0) * 100  # to %
        result["commission_rate"] = (ws1["AQ26"].value or 0) * 100

    wb.close()
    return result


def _apply_quote_to_session(q: dict) -> None:
    """Write parsed quote data into session_state so form fields pick it up.

    Key naming convention for _equipment_selector manual mode:
      - {prefix}_input_mode_{i} = "手動入力"
      - {prefix}_model_{i}       = model text
      - {prefix}_manual_output_{i} = output value (W/kW/kWh)
      - {prefix}_manual_count_{i}  = count
    """
    import re as _re_mod

    # Panel data → switch to manual input mode
    panels = q.get("panels", [])
    if panels:
        st.session_state["panel_types"] = len(panels)
        st.session_state["panel_input_mode"] = "手動入力"  # radio key (no suffix)
        for i, p in enumerate(panels):
            w_match = _re_mod.search(r"(\d+)", str(p.get("output", "")))
            watt = float(w_match.group(1)) if w_match else 0
            st.session_state[f"panel_model_{i}"] = f"{p['maker']} {p['model']}".strip()
            st.session_state[f"panel_manual_output_{i}"] = watt
            st.session_state[f"panel_manual_count_{i}"] = p["count"]

    # PCS data
    pcs_list = q.get("pcs_list", [])
    if pcs_list:
        st.session_state["pcs_types"] = len(pcs_list)
        st.session_state["pcs_input_mode"] = "手動入力"
        for i, pc in enumerate(pcs_list):
            kw_match = _re_mod.search(r"(\d+)", str(pc.get("output", "")))
            kw = float(kw_match.group(1)) if kw_match else 0
            st.session_state[f"pcs_model_{i}"] = f"{pc['maker']} {pc['model']}".strip()
            st.session_state[f"pcs_manual_output_{i}"] = kw
            st.session_state[f"pcs_manual_count_{i}"] = pc["count"]

    # Battery data
    bats = q.get("batteries", [])
    if bats:
        st.session_state["bat_types"] = len(bats)
        st.session_state["bat_input_mode"] = "手動入力"
        for i, b in enumerate(bats):
            st.session_state[f"bat_model_{i}"] = f"{b['maker']} {b['model']}".strip()
            st.session_state[f"bat_manual_output_{i}"] = b["kwh"]
            st.session_state[f"bat_manual_count_{i}"] = 1

    # Pricing data
    st.session_state["_quote_kw_unit_cost"] = int(round(q.get("kw_unit_cost", 0)))
    st.session_state["_quote_gross_margin_pct"] = round(q.get("gross_margin_pct", 0), 1)
    st.session_state["_quote_commission_rate"] = round(q.get("commission_rate", 0), 1)
    # Store exact selling price from quote to avoid rounding differences
    st.session_state["_quote_selling_price"] = int(q.get("selling_price", 0))
    st.session_state["_quote_raw_cost"] = int(q.get("raw_cost", 0))


def load_electricity_master() -> list[dict]:
    """Load contract electricity rates from Excel 契約電力マスタ sheet.

    Returns list of dicts with keys: company, contract, basic, peak, summer, other, night.
    Rates are 税込 (new unit prices, cols P-T = 16-20).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
        ws = wb["契約電力マスタ"]
        records = []
        for row in ws.iter_rows(min_row=5, max_col=20, values_only=False):
            company = row[0].value  # col A
            contract = row[1].value  # col B
            if not company or company == "電力会社":
                continue
            records.append({
                "company": str(company),
                "contract": str(contract),
                "basic": row[15].value or 0,     # col P (16th, 0-indexed=15)
                "peak": row[16].value,            # col Q
                "summer": row[17].value or 0,     # col R
                "other": row[18].value or 0,      # col S
                "night": row[19].value,           # col T
            })
        wb.close()
        return records
    except Exception:
        return []


@st.cache_data(ttl=3600)
def load_co2_factors() -> list[dict]:
    """Load CO2 emission factors from Excel CO2計算 sheet.

    Returns list of dicts with keys: name, real, adj.
    - name: 電気事業者名
    - real: 実排出係数 [tCO2/kWh] (col E)
    - adj:  調整後排出係数 [tCO2/kWh] (col J)
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
        co2_sheet = None
        for sn in wb.sheetnames:
            if "CO2" in sn and len(sn) < 10:
                co2_sheet = sn
                break
        if not co2_sheet:
            wb.close()
            return []
        ws = wb[co2_sheet]
        records = []
        for row in ws.iter_rows(min_row=7, max_col=13, values_only=False):
            name = row[0].value  # col A
            real_factor = row[4].value  # col E
            adj_factor = row[9].value   # col J
            if name and isinstance(real_factor, (int, float)):
                records.append({
                    "name": str(name).strip(),
                    "real": real_factor,
                    "adj": adj_factor if isinstance(adj_factor, (int, float)) else real_factor,
                })
        wb.close()
        return records
    except Exception:
        return []


def load_equipment_master() -> tuple[dict[str, list[dict]], str]:
    """Load active equipment records from Salesforce, grouped by MachineType__c.

    Returns (grouped_records, error_message). error_message is empty on success.
    """
    query = (
        "SELECT Name, MachineType__c, Maker__c, Katasiki__c, Output__c "
        "FROM EquipmentMaster__c "
        "WHERE Field1__c = false "
        "ORDER BY Maker__c, Katasiki__c"
    )
    try:
        records = _sf_query(query, timeout=30)
    except Exception as e:
        return {}, str(e)
    if not records:
        return {}, "No records returned"
    grouped: dict[str, list[dict]] = {}
    for r in records:
        mtype = r.get("MachineType__c", "")
        maker = r.get("Maker__c")
        katasiki = r.get("Katasiki__c")
        if not maker or not katasiki:
            continue
        grouped.setdefault(mtype, []).append({
            "name": r.get("Name", ""),
            "maker": maker,
            "katasiki": katasiki,
            "output": r.get("Output__c") or 0.0,
        })
    return grouped, ""


@st.cache_data(ttl=300, show_spinner="Salesforce検索中...")
def sf_search_opportunities(keyword: str) -> list[dict]:
    """Search Salesforce opportunities by name. Cached 5min."""
    safe_kw = keyword.replace("'", "").replace('"', "")
    tokens = _tokenize_keyword(safe_kw)
    like_clauses = " AND ".join(f"Name LIKE '%{t}%'" for t in tokens)
    query = (
        f"SELECT Id, Name, AccountId, Account.Name, "
        f"Account.BillingStreet, Account.BillingCity, Account.BillingState "
        f"FROM Opportunity WHERE {like_clauses} "
        f"ORDER BY LastModifiedDate DESC LIMIT 20"
    )
    try:
        return _sf_query(query)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Session state defaults
# ---------------------------------------------------------------------------

def _init_state():
    """Ensure session state keys exist."""
    defaults = {
        "panel_types": 1,
        "pcs_types": 1,
        "battery_types": 1,
        "sf_company": "",
        "sf_office": "",
        "sf_address": "",
        "sf_opp_id": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _round_100(val: float) -> int:
    """Round to nearest 100 yen."""
    return int(round(val / 100) * 100)

_init_state()

# ---------------------------------------------------------------------------
# Equipment master (load once)
# ---------------------------------------------------------------------------
def _get_eq_master() -> tuple[dict[str, list[dict]], str]:
    """Lazy-load equipment master (called inside tabs, not at module level)."""
    cached = st.session_state.get("eq_master_cache")
    # Retry if previous attempt failed (don't cache errors permanently)
    if cached is None or (cached[1] and not cached[0]):
        data, err = load_equipment_master()
        st.session_state["eq_master_cache"] = (data, err)
    return st.session_state["eq_master_cache"]


def _get_makers(eq_master: dict, machine_type: str) -> list[str]:
    """Return sorted unique maker list for a machine type."""
    records = eq_master.get(machine_type, [])
    return sorted(set(r["maker"] for r in records))


def _get_models(eq_master: dict, machine_type: str, maker: str) -> list[dict]:
    """Return model records for a given machine type + maker."""
    return [r for r in eq_master.get(machine_type, []) if r["maker"] == maker]


def _equipment_selector(
    machine_type: str,
    key_prefix: str,
    unit_label: str,
    output_multiplier: float,
    type_count_key: str,
    max_types: int = 3,
) -> tuple[list[dict], float, int]:
    """Render equipment input rows with cascading SF selectbox or manual input.

    Returns (data_list, total_value, total_count).
    total_value is in kW for panels/PCS, kWh for batteries.
    """
    eq_master, eq_error = _get_eq_master()
    eq_loaded = len(eq_master) > 0
    makers = _get_makers(eq_master, machine_type)
    sf_available = eq_loaded and len(makers) > 0

    # Input mode toggle
    if sf_available:
        input_mode = st.radio(
            "入力方式",
            ["Salesforceマスタ", "手動入力"],
            key=f"{key_prefix}_input_mode",
            horizontal=True,
        )
    else:
        input_mode = "手動入力"
        if not eq_loaded:
            msg = "⚠️ Salesforce接続不可 — 手動入力モードで動作中"
            if eq_error:
                msg += f"\n\nエラー: `{eq_error}`"
            st.caption(msg)

    data_list: list[dict] = []

    for i in range(st.session_state[type_count_key]):
        n = st.session_state[type_count_key]
        label_suffix = f" {i + 1}種類目" if n > 1 else ""
        if i > 0:
            st.markdown("---")

        if input_mode == "Salesforceマスタ":
            c1, c2, c3, c4, c5 = st.columns([1.5, 2, 1, 1, 1])
            with c1:
                sel_maker = st.selectbox(
                    f"メーカー{label_suffix}",
                    [""] + makers,
                    key=f"{key_prefix}_maker_{i}",
                )
            with c2:
                if sel_maker:
                    models = _get_models(eq_master, machine_type, sel_maker)
                    model_options = [""] + [m["katasiki"] for m in models]
                else:
                    models = []
                    model_options = [""]
                sel_katasiki = st.selectbox(
                    f"型式{label_suffix}",
                    model_options,
                    key=f"{key_prefix}_katasiki_{i}",
                )
            # Auto-populate output
            auto_output = 0.0
            if sel_maker and sel_katasiki:
                match = next((m for m in models if m["katasiki"] == sel_katasiki), None)
                if match:
                    auto_output = match["output"] * output_multiplier
            with c3:
                _col3_label = "容量 (kWh)" if unit_label == "kWh" else f"出力 ({unit_label})"
                if unit_label == "W" and auto_output > 0:
                    st.metric(_col3_label, f"{auto_output:.0f}")
                elif auto_output > 0:
                    st.metric(_col3_label, f"{auto_output:.1f}")
                else:
                    st.metric(_col3_label, "—")
                unit_val = auto_output
            with c4:
                _sf_cnt_key = f"{key_prefix}_count_{i}"
                if _sf_cnt_key not in st.session_state:
                    st.session_state[_sf_cnt_key] = 0
                count = st.number_input(
                    "枚数" if unit_label == "W" else "台数",
                    min_value=0, step=1,
                    key=_sf_cnt_key,
                )
            with c5:
                if unit_label == "W":
                    total = unit_val * count / 1000.0
                    st.metric("小計 (kW)", f"{total:,.2f}")
                elif unit_label == "kWh":
                    total = unit_val * count
                    st.metric("小計 (kWh)", f"{total:,.1f}")
                else:
                    total = unit_val * count
                    st.metric("小計 (kW)", f"{total:,.2f}")

            model_str = f"{sel_maker} {sel_katasiki}" if sel_maker and sel_katasiki else ""

        else:
            # Manual input mode
            _placeholders = {
                "W": "例：カナディアンソーラー CS6R-400MS",
                "kW": "例：Huawei SUN2000-50KTL-JPM1",
                "kWh": "例：Huawei LUNA2000-21-S0",
            }
            c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
            with c1:
                model_str = st.text_input(
                    f"メーカー・型番{label_suffix}",
                    placeholder=_placeholders.get(unit_label, ""),
                    key=f"{key_prefix}_model_{i}",
                )
            with c2:
                _out_key = f"{key_prefix}_manual_output_{i}"
                if _out_key not in st.session_state:
                    st.session_state[_out_key] = 400.0 if unit_label == "W" else 0.0
                unit_val = st.number_input(
                    f"1枚あたり出力 ({unit_label})" if unit_label == "W"
                    else f"1台あたり{'容量' if unit_label == 'kWh' else '出力'} ({unit_label})",
                    min_value=0.0, step=10.0 if unit_label == "W" else 0.5,
                    key=_out_key,
                )
            with c3:
                _cnt_key = f"{key_prefix}_manual_count_{i}"
                if _cnt_key not in st.session_state:
                    st.session_state[_cnt_key] = 0
                count = st.number_input(
                    "枚数" if unit_label == "W" else "台数",
                    min_value=0, step=1,
                    key=_cnt_key,
                )
            with c4:
                if unit_label == "W":
                    total = unit_val * count / 1000.0
                    st.metric("小計 (kW)", f"{total:,.2f}")
                elif unit_label == "kWh":
                    total = unit_val * count
                    st.metric("小計 (kWh)", f"{total:,.1f}")
                else:
                    total = unit_val * count
                    st.metric("小計 (kW)", f"{total:,.2f}")

        # Build data dict (compatible with existing structure)
        entry: dict = {
            "type_index": i + 1,
            "model": model_str,
            "count": count,
        }
        if unit_label == "W":
            entry["watt_per_unit"] = unit_val
            entry["total_kw"] = total
        elif unit_label == "kWh":
            entry["kwh_per_unit"] = unit_val
            entry["total_kwh"] = total
        else:  # kW
            entry["kw_per_unit"] = unit_val
            entry["total_kw"] = total
        data_list.append(entry)

    # Add / remove buttons
    btn_a, btn_b, _ = st.columns([1, 1, 4])
    with btn_a:
        if st.session_state[type_count_key] < max_types:
            if st.button("＋ 種類を追加", key=f"add_{key_prefix}"):
                st.session_state[type_count_key] += 1
                st.rerun()
    with btn_b:
        if st.session_state[type_count_key] > 1:
            if st.button("－ 最後を削除", key=f"rm_{key_prefix}"):
                st.session_state[type_count_key] -= 1
                st.rerun()

    # Totals
    if unit_label == "kWh":
        total_val = sum(d["total_kwh"] for d in data_list)
    else:
        total_val = sum(d["total_kw"] for d in data_list)
    total_cnt = sum(d["count"] for d in data_list)
    return data_list, total_val, total_cnt

# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------

if LOGO_PATH.exists():
    st.image(str(LOGO_PATH), width=180)
st.title("PPA/EPC 提案資料ジェネレーター")
st.caption("変数を入力してスライド構成を選択し、PPTX を生成します")

tab1, tab2, tab3, tab4 = st.tabs([
    "① 顧客情報",
    "② 案件詳細",
    "③ スライド構成",
    "④ 生成・出力",
])

# =========================================================================
# Tab 1: Customer Info
# =========================================================================

with tab1:
    with st.expander("💾 案件データ 保存・読込", expanded=False):
        save_col, load_col = st.columns(2)

        with save_col:
            st.markdown("**保存**")
            if st.button("現在の入力データを保存", key="save_case"):
                _cdata = st.session_state.get("customer_data", {})
                if _cdata and _cdata.get("company_name"):
                    _company = _cdata.get("company_name", "unknown")
                    _ptype = _cdata.get("proposal_type", "ppa")
                    _date = _cdata.get("proposal_date", "")
                    _fname = f"{_company}_{_ptype}_{_date}.json"
                    _fname = _re.sub(r'[\\/*?:"<>|]', '_', _fname)
                    _fpath = SAVE_DIR / _fname
                    with open(_fpath, "w", encoding="utf-8") as _f:
                        json.dump(_cdata, _f, ensure_ascii=False, indent=2, default=str)
                    st.success(f"保存しました: {_fname}")
                else:
                    st.warning("顧客情報を入力してから保存してください")
            if st.session_state.get("customer_data"):
                _dl_json = json.dumps(
                    st.session_state["customer_data"],
                    ensure_ascii=False, indent=2, default=str,
                )
                st.download_button(
                    "📥 JSONダウンロード",
                    data=_dl_json,
                    file_name=f"case_{st.session_state.get('customer_data', {}).get('company_name', 'data')}.json",
                    mime="application/json",
                )

        with load_col:
            st.markdown("**読込**")
            _saved = sorted(SAVE_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
            if _saved:
                _opts = [""] + [f.name for f in _saved]
                _sel = st.selectbox("保存済み案件を選択", _opts, key="load_case_select")
                if _sel and st.button("読み込む", key="load_case"):
                    with open(SAVE_DIR / _sel, "r", encoding="utf-8") as _f:
                        _loaded = json.load(_f)
                    st.session_state["customer_data"] = _loaded
                    st.session_state["sf_company"] = _loaded.get("company_name", "")
                    st.session_state["sf_office"] = _loaded.get("office_name", "")
                    st.session_state["sf_address"] = _loaded.get("address", "")
                    st.success(f"読み込みました: {_sel}")
                    st.rerun()
            else:
                st.info("保存済みの案件はありません")

            _upload = st.file_uploader("またはJSONをアップロード", type=["json"], key="upload_case")
            if _upload:
                _loaded = json.load(_upload)
                st.session_state["customer_data"] = _loaded
                st.session_state["sf_company"] = _loaded.get("company_name", "")
                st.session_state["sf_office"] = _loaded.get("office_name", "")
                st.session_state["sf_address"] = _loaded.get("address", "")
                st.success("アップロードしたデータを読み込みました")
                st.rerun()

        # ----- Box integration -----
        from proposal_generator.box_client import is_available as _box_ok

        if _box_ok():
            st.divider()
            st.markdown("**📦 Box連携（03_提案資料）**")
            _box_deal = st.text_input(
                "商談名で検索", key="box_deal_search",
                value=st.session_state.get("sf_office", ""),
                placeholder="Boxフォルダ名（商談名）を入力",
            )
            if _box_deal and st.button("Boxフォルダを検索", key="box_search_btn"):
                try:
                    from proposal_generator.box_client import get_deal_proposal_folder, list_files
                    _folder_id = get_deal_proposal_folder(_box_deal)
                    if _folder_id:
                        st.session_state["box_proposal_folder_id"] = _folder_id
                        _files = list_files(_folder_id)
                        st.session_state["box_file_list"] = _files
                        st.success(f"03_提案資料フォルダを発見 ({len(_files)} ファイル)")
                    else:
                        st.warning(f"「{_box_deal}」に一致する商談フォルダが見つかりません")
                except Exception as e:
                    st.error(f"Box検索エラー: {e}")

            _box_files = st.session_state.get("box_file_list", [])
            if _box_files:
                _json_files = [f for f in _box_files if f["name"].endswith(".json")]
                if _json_files:
                    _box_sel = st.selectbox(
                        "Boxから案件JSONを読込",
                        [""] + [f["name"] for f in _json_files],
                        key="box_load_select",
                    )
                    if _box_sel and st.button("Boxから読み込む", key="box_load_btn"):
                        try:
                            from proposal_generator.box_client import download_file
                            _fid = next(f["id"] for f in _json_files if f["name"] == _box_sel)
                            _tmp = Path(tempfile.mktemp(suffix=".json"))
                            download_file(_fid, _tmp)
                            with open(_tmp, "r", encoding="utf-8") as _bf:
                                _loaded = json.load(_bf)
                            _tmp.unlink(missing_ok=True)
                            st.session_state["customer_data"] = _loaded
                            st.session_state["sf_company"] = _loaded.get("company_name", "")
                            st.session_state["sf_office"] = _loaded.get("office_name", "")
                            st.session_state["sf_address"] = _loaded.get("address", "")
                            st.success(f"Boxから読み込みました: {_box_sel}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Box読込エラー: {e}")

            # Upload JSON to Box
            if st.session_state.get("customer_data") and st.session_state.get("box_proposal_folder_id"):
                if st.button("案件JSONをBoxに保存", key="box_save_json_btn"):
                    try:
                        from proposal_generator.box_client import upload_file as _box_upload
                        _cdata = st.session_state["customer_data"]
                        _company = _cdata.get("company_name", "unknown")
                        _ptype = _cdata.get("proposal_type", "ppa")
                        _date = _cdata.get("proposal_date", "")
                        _fname = f"{_company}_{_ptype}_{_date}.json"
                        _fname = _re.sub(r'[\\/*?:"<>|]', '_', _fname)
                        _tmp = Path(tempfile.mktemp(suffix=".json"))
                        with open(_tmp, "w", encoding="utf-8") as _bf:
                            json.dump(_cdata, _bf, ensure_ascii=False, indent=2, default=str)
                        _result = _box_upload(
                            st.session_state["box_proposal_folder_id"], _tmp, _fname
                        )
                        _tmp.unlink(missing_ok=True)
                        st.success(f"Boxに保存しました: {_result['name']}")
                    except Exception as e:
                        st.error(f"Box保存エラー: {e}")
        else:
            st.caption("📦 Box連携: box_config.json未設定 (access_tokenを設定するとBoxフォルダと連携できます)")

    with st.expander("🔍 Salesforceから取引先・商談を検索", expanded=True):
        sf_keyword = st.text_input(
            "商談名で検索（Enter で実行）",
            placeholder="例：田中貴金属、Mizkan、コスモ精機 など",
            key="sf_keyword",
        )

        if sf_keyword:
            opp_records = sf_search_opportunities(sf_keyword)
            if opp_records:
                opp_options = [
                    f"{r['Name']}  ／  {r.get('Account', {}).get('Name', '—')}"
                    for r in opp_records
                ]
                selected_idx = st.selectbox(
                    "商談を選択（選択すると自動入力されます）",
                    options=range(len(opp_records)),
                    format_func=lambda i: opp_options[i],
                    key="sf_selected_idx",
                )
                sel = opp_records[selected_idx]
                acct = sel.get("Account") or {}
                parts = [
                    acct.get("BillingState", ""),
                    acct.get("BillingCity", ""),
                    acct.get("BillingStreet", ""),
                ]
                st.session_state["sf_company"] = acct.get("Name", "")
                st.session_state["sf_office"] = sel.get("Name", "")
                st.session_state["sf_address"] = "".join(p for p in parts if p)
                st.session_state["sf_opp_id"] = sel.get("Id", "")
            else:
                st.info("該当する商談が見つかりませんでした")

    st.divider()
    st.subheader("顧客・案件情報")

    company_name = st.session_state.get("sf_company", "")
    office_name = st.session_state.get("sf_office", "")
    address = st.session_state.get("sf_address", "")

    sf_col_a, sf_col_b, sf_col_c = st.columns(3)
    with sf_col_a:
        st.markdown("**取引先名**")
        if company_name:
            st.markdown(f"#### {company_name}")
        else:
            st.markdown("*未選択*")
    with sf_col_b:
        st.markdown("**商談**")
        if office_name:
            st.markdown(f"#### {office_name}")
        else:
            st.markdown("*未選択*")
    with sf_col_c:
        st.markdown("**所在地**")
        if address:
            st.markdown(f"#### {address}")
        else:
            st.markdown("*未選択*")

    if not company_name:
        st.caption("⬆ 上の検索から取引先・商談を選択してください")

    st.divider()

    col1, col2, col3 = st.columns(3)
    with col1:
        company_size = st.selectbox(
            "企業規模",
            ["", "大企業", "中小企業", "その他（学校法人・医療法人等）"],
        )
    with col2:
        proposal_date = st.date_input("提案日")
        site_survey = st.selectbox("現地調査", ["", "実施済み", "未実施"])
    with col3:
        tax_display = st.selectbox("提案書税表記", ["税抜", "税込"])
        snow_depth = st.number_input("垂直積雪量 (cm)", min_value=0, step=10, value=0)

# =========================================================================
# Tab 2: Project Details
# =========================================================================

with tab2:
    # ----- Proposal Type -----
    proposal_type = st.radio(
        "提案タイプ",
        ["PPA（第三者所有）", "EPC（販売）"],
        key="proposal_type",
        horizontal=True,
    )
    is_epc = proposal_type.startswith("EPC")

    # ----- Quote Import (top of Tab 2) -----
    with st.expander("📄 見積書から一括読み込み", expanded=False):
        _quote_file = st.file_uploader(
            "見積書Excelをアップロード（.xlsm / .xlsx）",
            type=["xlsm", "xlsx", "xls"],
            key="quote_file",
            help="当社標準テンプレートの見積書から機器・価格情報を自動入力",
        )
        if _quote_file is not None:
            try:
                _q = _parse_quote_excel(_quote_file)
                if _q:
                    st.session_state["quote_data"] = _q
                    st.success(f"読み込み完了: パネル{_q.get('panel_count', 0):,}枚 / PCS{_q.get('pcs_count', 0):,}台")
                    with st.expander("読み込み内容を確認", expanded=False):
                        st.json(_q)
                    if st.button("📥 この内容をフォームに反映する", key="apply_quote", type="primary"):
                        _apply_quote_to_session(_q)
                        st.success("反映しました。各項目を確認してください。")
                        st.rerun()
            except Exception as e:
                st.error(f"読み込みエラー: {e}")

    st.divider()

    # ----- Panel -----
    with st.expander("🔆 パネル情報", expanded=True):
        panel_data_list, total_panel_kw, total_panel_count = _equipment_selector(
            machine_type="モジュール",
            key_prefix="panel",
            unit_label="W",
            output_multiplier=1000.0,  # SF stores kW, display as W
            type_count_key="panel_types",
        )
        st.info(f"パネル合計: **{total_panel_count:,}枚** / **{total_panel_kw:,.2f} kW**")

    # ----- PCS -----
    with st.expander("⚡ パワコン（PCS）情報", expanded=True):
        pcs_data_list, total_pcs_kw, total_pcs_count = _equipment_selector(
            machine_type="PCS",
            key_prefix="pcs",
            unit_label="kW",
            output_multiplier=1.0,
            type_count_key="pcs_types",
        )
        st.info(f"PCS合計: **{total_pcs_count:,}台** / **{total_pcs_kw:,.2f} kW**")

    # ----- Battery -----
    with st.expander("🔋 蓄電池情報", expanded=True):
        battery_data_list, total_battery_kwh, total_battery_count = _equipment_selector(
            machine_type="蓄電池",
            key_prefix="bat",
            unit_label="kWh",
            output_multiplier=1.0,
            type_count_key="battery_types",
        )
        if total_battery_count > 0:
            st.info(f"蓄電池合計: **{total_battery_count:,}台** / **{total_battery_kwh:,.1f} kWh**")

    # Auto-calc system capacity (= min of panel kW and PCS kW)
    if total_pcs_kw > 0:
        system_capacity = round(min(total_panel_kw, total_pcs_kw), 2)
    else:
        system_capacity = total_panel_kw

    # ----- Contract Terms -----
    with st.expander("📋 契約条件", expanded=True):
        _auto_demand_cut = st.session_state.get("ipals_data", {}).get("demand_cut_kw", 0.0) or 0.0
        if is_epc:
            ct_col1, ct_col2 = st.columns(2)
            with ct_col1:
                contract_years = st.number_input("契約期間 (年)", min_value=1, max_value=30, value=20)
            with ct_col2:
                demand_reduction = st.number_input(
                    "削減デマンド (kW)", min_value=0.0, step=1.0, value=_auto_demand_cut,
                    help="iPals CSVアップロード時は自動算出されます",
                )
            ppa_unit_price = 0.0
            surplus_price = 0.0
        else:
            ct_col1, ct_col2, ct_col3, ct_col4 = st.columns(4)
            with ct_col1:
                contract_years = st.number_input("契約期間 (年)", min_value=1, max_value=30, value=20)
            with ct_col2:
                ppa_unit_price = st.number_input(
                    "PPA単価 (円/kWh)",
                    min_value=0.0, step=0.5,
                    key="ppa_price_input",
                    help="下の「PPA単価自動試算」ボタンで自動入力できます",
                )
            with ct_col3:
                surplus_price = st.number_input("余剰売電単価 (円/kWh)", min_value=0.0, step=0.5)
            with ct_col4:
                demand_reduction = st.number_input(
                    "削減デマンド (kW)", min_value=0.0, step=1.0, value=_auto_demand_cut,
                    help="iPals CSVアップロード時は自動算出されます",
                )

        # ----- Current Electricity Cost (contract master based) -----
        st.markdown("**現在の電気料金**")
        _elec_master = load_electricity_master()
        _co2_factors = load_co2_factors()
        _co2_factor_val = 0.000453  # default: national average
        if _elec_master:
            _companies = sorted(set(r["company"] for r in _elec_master))
            _companies_with_manual = _companies + ["その他（新電力・手入力）"]
            _elec_company = st.selectbox("電力会社", [""] + _companies_with_manual, key="elec_company")

            if _elec_company and _elec_company != "その他（新電力・手入力）":
                _contracts = [r for r in _elec_master if r["company"] == _elec_company]
                _contract_names = [r["contract"] for r in _contracts]
                _elec_contract = st.selectbox("契約種別", [""] + _contract_names, key="elec_contract")

                if _elec_contract:
                    _sel = next((r for r in _contracts if r["contract"] == _elec_contract), None)
                    if _sel:
                        # Editable fee inputs with master defaults
                        _master_basic = float(_sel["basic"])
                        _master_summer = float(_sel["summer"] or _sel["other"] or 0)
                        _master_other = float(_sel["other"] or _sel["summer"] or 0)

                        _fc1, _fc2, _fc3 = st.columns(3)
                        with _fc1:
                            _basic_rate = st.number_input(
                                "基本料金単価 (円/kW)", min_value=0.0, step=10.0,
                                value=_master_basic, key="elec_basic_rate",
                                help=f"マスタ値: ¥{_master_basic:,.1f}/kW")
                        with _fc2:
                            _summer_rate = st.number_input(
                                "夏季従量単価 (円/kWh)", min_value=0.0, step=0.5,
                                value=_master_summer, key="elec_summer_rate",
                                help=f"マスタ値: ¥{_master_summer:.2f}/kWh")
                        with _fc3:
                            _other_rate = st.number_input(
                                "その他季従量単価 (円/kWh)", min_value=0.0, step=0.5,
                                value=_master_other, key="elec_other_rate",
                                help=f"マスタ値: ¥{_master_other:.2f}/kWh")

                        st.session_state["_basic_rate_kw"] = _basic_rate

                        _ep1, _ep2, _ep3 = st.columns([2, 2, 1])
                        with _ep1:
                            _contract_kw = st.number_input("契約電力 (kW)", min_value=0.0, step=1.0, key="contract_kw")
                        with _ep2:
                            _annual_kwh = st.number_input("年間使用電力量 (kWh)", min_value=0, step=1000, key="annual_kwh")
                        with _ep3:
                            st.number_input("力率 (%)", min_value=50, max_value=100, value=85, step=5, key="power_factor_pct",
                                            help="一般的な高圧受電は85%です")

                        # Calculate annual cost
                        _basic_annual = _basic_rate * _contract_kw * 12
                        _avg_unit = (_summer_rate * 4 + _other_rate * 8) / 12
                        _usage_annual = _avg_unit * _annual_kwh
                        annual_elec_cost = int(_basic_annual + _usage_annual)
                        st.caption(
                            f"年間電気代（概算）: **¥{annual_elec_cost:,.0f}**　"
                            f"（基本: ¥{_basic_annual:,.0f} + 従量: ¥{_usage_annual:,.0f}　"
                            f"加重平均単価: ¥{_avg_unit:.2f}/kWh）"
                        )

                        # CO2 emission factor: auto-match from CO2計算 sheet
                        _co2_match = next((f for f in _co2_factors if f["name"] == _elec_company), None)
                        if _co2_match:
                            _co2_factor_val = _co2_match["adj"]
                            st.caption(f"CO₂排出係数（調整後）: **{_co2_factor_val:.6f}** tCO₂/kWh — {_elec_company}")
                    else:
                        annual_elec_cost = 0
                else:
                    annual_elec_cost = 0
            elif _elec_company == "その他（新電力・手入力）":
                _mc1, _mc2, _mc3 = st.columns(3)
                with _mc1:
                    _manual_basic = st.number_input("基本料金単価 (円/kW)", min_value=0.0, step=100.0, key="manual_basic")
                with _mc2:
                    _manual_rate = st.number_input("従量単価 (円/kWh)", min_value=0.0, step=0.5, key="manual_rate")
                with _mc3:
                    _manual_kw = st.number_input("契約電力 (kW)", min_value=0.0, step=1.0, key="manual_contract_kw")

                _mr1, _mr2 = st.columns(2)
                with _mr1:
                    _manual_kwh = st.number_input("年間使用電力量 (kWh)", min_value=0, step=1000, key="manual_annual_kwh")
                with _mr2:
                    st.number_input("力率 (%)", min_value=50, max_value=100, value=85, step=5, key="power_factor_pct",
                                    help="一般的な高圧受電は85%です")

                st.session_state["_basic_rate_kw"] = _manual_basic

                _basic_annual = _manual_basic * _manual_kw * 12
                _usage_annual = _manual_rate * _manual_kwh
                annual_elec_cost = int(_basic_annual + _usage_annual)
                if annual_elec_cost > 0:
                    st.caption(
                        f"年間電気代（概算）: **¥{annual_elec_cost:,.0f}**　"
                        f"（基本: ¥{_basic_annual:,.0f} + 従量: ¥{_usage_annual:,.0f}）"
                    )

                # CO2 emission factor: select from CO2計算 sheet
                if _co2_factors:
                    _co2_names = [""] + [f["name"] for f in _co2_factors]
                    _co2_selected = st.selectbox(
                        "CO₂排出係数 — 電気事業者", _co2_names,
                        key="co2_company",
                        help="CO2計算シートから事業者を選択（調整後排出係数を使用）",
                    )
                    if _co2_selected:
                        _co2_match = next((f for f in _co2_factors if f["name"] == _co2_selected), None)
                        if _co2_match:
                            _co2_factor_val = _co2_match["adj"]
                            st.caption(f"調整後排出係数: **{_co2_factor_val:.6f}** tCO₂/kWh")
            else:
                annual_elec_cost = 0
        else:
            st.caption("⚠️ 契約電力マスタを読み込めません（Excelファイル未設定）")
            annual_elec_cost = st.number_input(
                "現在の年間電気代 (円)", min_value=0, step=100000, value=0,
                help="PP7・PP8の電気代削減額試算に使用します",
            )

        # Store CO2 emission factor in session state
        st.session_state["_co2_emission_factor"] = _co2_factor_val

    # ----- Pricing -----
    with st.expander("💰 価格情報", expanded=False):

        price_col1, price_col2 = st.columns(2)
        with price_col1:
            st.markdown("**入力項目**")
            # Use quote data as defaults if available
            if "_quote_kw_unit_cost" in st.session_state and "kw_unit_cost_input" not in st.session_state:
                st.session_state["kw_unit_cost_input"] = st.session_state["_quote_kw_unit_cost"]
            if "_quote_gross_margin_pct" in st.session_state and "gross_margin_input" not in st.session_state:
                st.session_state["gross_margin_input"] = st.session_state["_quote_gross_margin_pct"]
            if "_quote_commission_rate" in st.session_state and "commission_input" not in st.session_state:
                st.session_state["commission_input"] = st.session_state["_quote_commission_rate"]

            if "kw_unit_cost_input" not in st.session_state:
                st.session_state["kw_unit_cost_input"] = 0
            if "gross_margin_input" not in st.session_state:
                st.session_state["gross_margin_input"] = 0.0
            if "commission_input" not in st.session_state:
                st.session_state["commission_input"] = 0.0

            kw_unit_cost = st.number_input(
                "kW原価 (円/kW)",
                min_value=0, step=1000,
                key="kw_unit_cost_input",
                help="設備1kWあたりの原価",
            )
            gross_margin_pct = st.number_input(
                "粗利率 (%)",
                min_value=0.0, max_value=99.9, step=0.5,
                key="gross_margin_input",
            )
            sales_commission_pct = st.number_input(
                "販売手数料 (%)",
                min_value=0.0, max_value=100.0, step=0.5,
                key="commission_input",
                help="例：販売価格の3%を手数料として支払う場合",
            )

        with price_col2:
            st.markdown("**自動算出**")
            # If quote data has exact prices, use those
            _q_sell = st.session_state.get("_quote_selling_price", 0)
            _q_raw = st.session_state.get("_quote_raw_cost", 0)

            if _q_raw > 0 and _q_sell > 0:
                # Use exact values from quote
                raw_cost = _q_raw
                selling_price = _q_sell
                gross_profit = selling_price - raw_cost
            else:
                # Calculate from kW unit cost and margin
                raw_cost = kw_unit_cost * total_panel_kw
                margin_rate = gross_margin_pct / 100.0
                if margin_rate > 0 and margin_rate < 1 and raw_cost > 0:
                    gross_profit = raw_cost * margin_rate / (1 - margin_rate)
                    selling_price_raw = raw_cost + gross_profit
                    selling_price = _round_100(selling_price_raw)
                else:
                    gross_profit = 0.0
                    selling_price = 0

            commission_amount = (
                selling_price * sales_commission_pct / 100.0
                if sales_commission_pct > 0 and selling_price > 0
                else 0.0
            )

            st.metric(
                "原価",
                f"¥{raw_cost:,.0f}",
                help="見積書から取得" if _q_raw > 0 else f"kW原価 × パネル合計kW",
            )
            st.metric("粗利", f"¥{gross_profit:,.0f}")
            _kw_sell = int(selling_price / total_panel_kw) if selling_price and total_panel_kw > 0 else 0
            st.metric("販売価格", f"¥{selling_price:,.0f}",
                      delta=f"¥{_kw_sell:,}/kW" if _kw_sell else None, delta_color="off")
            if commission_amount > 0:
                st.metric(
                    f"販売手数料（{sales_commission_pct:.1f}%）",
                    f"¥{commission_amount:,.0f}",
                )
                # Net margin = gross profit - commission
                net_profit = gross_profit - commission_amount
                net_margin_pct = (net_profit / selling_price * 100) if selling_price > 0 else 0.0
                st.metric(
                    "実質粗利",
                    f"¥{net_profit:,.0f}",
                    delta=f"実質粗利率 {net_margin_pct:.1f}%",
                    delta_color="normal",
                )
            else:
                net_profit = gross_profit
                net_margin_pct = gross_margin_pct

    # ----- Subsidy -----
    with st.expander("🏦 補助金", expanded=False):
        # Fixed subsidy program list
        _SUBSIDY_PROGRAMS = [
            "なし",
            "環境省（ストレージパリティ）",
            "東京都",
            "神奈川県",
            "埼玉県",
            "群馬県",
            "静岡県",
            "山梨県",
            "手動入力",
        ]
        if is_epc:
            _SUBSIDY_PROGRAMS.insert(5, "埼玉県（CO2排出削減・EPC専用）")

        _sub_sel = st.selectbox("補助金プログラム", _SUBSIDY_PROGRAMS, key="subsidy_select")

        # Auto-calc if program selected and data available
        _subsidy_results: dict[str, dict] = {}
        if total_panel_kw > 0 and total_pcs_kw > 0 and _sub_sel not in ("なし", "手動入力"):
            from proposal_generator.subsidy_calc import calc_all_subsidies

            _bat_price = st.number_input(
                "蓄電池販売価格 (円)", min_value=0, step=100000, value=0,
                key="battery_price_input",
                help="蓄電池の販売価格（補助金計算に使用）",
            )

            _all = calc_all_subsidies(
                panel_kw=total_panel_kw,
                pcs_kw=total_pcs_kw,
                selling_price=selling_price,
                battery_price=_bat_price,
                battery_kwh=total_battery_kwh,
                company_size=company_size,
                proposal_type="epc" if is_epc else "ppa",
            )
            _subsidy_results = {s["name"]: s for s in _all}

        if _sub_sel == "なし":
            subsidy_name = ""
            subsidy_amount = 0
        elif _sub_sel == "手動入力":
            subsidy_name = st.text_input("補助金名", placeholder="例：東京都補助金")
            subsidy_amount = st.number_input("補助金額 (円)", min_value=0, step=10000, value=0)
        elif _sub_sel == "山梨県":
            subsidy_name = "山梨県"
            subsidy_amount = st.number_input("補助金額 (円)（山梨県は手動入力）", min_value=0, step=10000, value=0)
            st.caption("山梨県の補助金計算は個別確認が必要です")
        else:
            # Map UI name to calc result name
            _name_map = {
                "環境省（ストレージパリティ）": "環境省（ストレージパリティ）",
                "東京都": "東京都",
                "神奈川県": "神奈川県",
                "埼玉県": "埼玉県②",
                "埼玉県（CO2排出削減・EPC専用）": "埼玉県①（CO2排出削減）",
                "群馬県": "群馬県",
                "静岡県": "静岡県",
            }
            _calc_name = _name_map.get(_sub_sel, _sub_sel)
            _match = _subsidy_results.get(_calc_name)

            subsidy_name = _sub_sel
            if _match and _match["applicable"]:
                subsidy_amount = _match["amount"]
                st.metric("自動試算額", f"¥{subsidy_amount:,.0f}")
                if _match.get("notes"):
                    st.caption(_match["notes"])
            elif _match and not _match["applicable"]:
                subsidy_amount = 0
                st.warning(_match.get("notes", "この補助金は現在の条件では適用できません"))
            else:
                subsidy_amount = 0
                st.info("パネル・PCS情報を入力すると自動試算します")

    # ----- Lease / Loan (PPA only) -----
    if not is_epc:
        with st.expander("📋 リース情報", expanded=False):
            _finance_companies = ["シーエナジー", "みずほリース", "群馬銀行"]
            lease_company = st.selectbox("調達先", _finance_companies, key="lease_company")

            _is_bank_loan = (lease_company == "群馬銀行")

            # Default rates and terms per company
            _finance_defaults = {
                "シーエナジー": {"rate": 3.10, "term": 20},
                "みずほリース": {"rate": 5.50, "term": 10},
                "群馬銀行": {"rate": 1.80, "term": 20},
            }
            _defaults = _finance_defaults[lease_company]

            # Labels switch based on finance type
            _company_label = "銀行" if _is_bank_loan else "リース会社"
            _term_label = "借入期間(年数)" if _is_bank_loan else "リース利用期間"
            _rate_label = "調達金利(年利)" if _is_bank_loan else "リース金利(年利)"

            st.info(f"**{_company_label}**: {lease_company}　／　デフォルト金利: **{_defaults['rate']:.2f}%**")

            lease_rate = st.number_input(
                _rate_label + " (%)", min_value=0.0, step=0.1,
                value=_defaults["rate"], key="lease_rate_input",
            )
            lease_years = st.number_input(
                _term_label, min_value=1, max_value=30,
                value=_defaults["term"], key="lease_years_input",
            )

            # Bank loan additional cost info
            if _is_bank_loan and selling_price > 0:
                _fire_insurance = math.ceil(selling_price / 1_000_000 * 3107 / 1000) * 1000
                st.caption(f"火災保険料: ¥{_fire_insurance:,.0f}/年　償却資産税率: 1.4%")
    else:
        lease_company = ""
        lease_rate = 0.0
        lease_years = 0

    # ----- iPals upload -----
    with st.expander("📊 iPals発電シミュレーション", expanded=False):
        ipals_file = st.file_uploader(
            "iPals出力CSVをアップロード（任意）",
            type=["csv"],
            help="iPals自家消費発電量CSVをアップロード → 年間発電量等を自動計算",
        )
        if ipals_file is not None:
            import csv
            import io
            # Try cp932 (Shift_JIS) first, then utf-8
            _raw = ipals_file.getvalue()
            for _enc in ("cp932", "utf-8-sig", "utf-8"):
                try:
                    _text = _raw.decode(_enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            else:
                _text = _raw.decode("utf-8", errors="replace")

            _reader = csv.reader(io.StringIO(_text))
            _headers = next(_reader, [])
            # Cols: 月,日,時,発電量(kWh),需要量(kWh),不足電力量(kWh),余剰電力量(kWh),自家消費電力量(kWh),自家消費率(%),消費率(%),モジュール出力(kWh)
            _total_gen = 0.0
            _total_demand = 0.0
            _total_surplus = 0.0
            _total_self_consume = 0.0
            _monthly_gen = [0.0] * 12
            _hourly_rows = []  # raw hourly data for demand calc / charts
            _row_count = 0
            for _row in _reader:
                if len(_row) < 8:
                    continue
                try:
                    _month = int(_row[0])
                    _day = int(_row[1])
                    _hour = int(_row[2])
                    _gen = float(_row[3]) if _row[3] and _row[3] != "-" else 0.0
                    _demand = float(_row[4]) if _row[4] and _row[4] != "-" else 0.0
                    _surplus = float(_row[6]) if _row[6] and _row[6] != "-" else 0.0
                    _self_c = float(_row[7]) if _row[7] and _row[7] != "-" else 0.0
                except (ValueError, IndexError):
                    continue
                _total_gen += _gen
                _total_demand += _demand
                _total_surplus += _surplus
                _total_self_consume += _self_c
                if 1 <= _month <= 12:
                    _monthly_gen[_month - 1] += _gen
                _hourly_rows.append({
                    "month": _month, "day": _day, "hour": _hour,
                    "demand_kw": _demand, "gen_kw": _gen,
                    "self_consumption_kw": _self_c, "surplus_kw": _surplus,
                })
                _row_count += 1

            if _row_count > 0:
                _self_rate = (_total_self_consume / _total_gen * 100) if _total_gen > 0 else 0.0
                _co2_ef = st.session_state.get("_co2_emission_factor", 0.000453)
                _co2_t = _total_gen * _co2_ef

                st.success(f"読み込み完了: {_row_count:,}行（{_row_count // 24}日分）")
                _ic1, _ic2, _ic3, _ic4 = st.columns(4)
                with _ic1:
                    st.metric("年間発電量", f"{_total_gen:,.0f} kWh")
                with _ic2:
                    st.metric("自家消費量", f"{_total_self_consume:,.0f} kWh")
                with _ic3:
                    st.metric("自家消費率", f"{_self_rate:.1f}%")
                with _ic4:
                    st.metric("余剰電力量", f"{_total_surplus:,.0f} kWh")

                # Peak demand detection
                _peak_before = max((r["demand_kw"] for r in _hourly_rows), default=0)
                _peak_after = max((r["demand_kw"] - r["self_consumption_kw"] for r in _hourly_rows), default=0)
                _demand_cut = _peak_before - _peak_after

                _dc1, _dc2, _dc3 = st.columns(3)
                with _dc1:
                    st.metric("ピークデマンド（導入前）", f"{_peak_before:,.0f} kW")
                with _dc2:
                    st.metric("ピークデマンド（導入後）", f"{_peak_after:,.0f} kW")
                with _dc3:
                    st.metric("▲デマンド削減", f"{_demand_cut:,.0f} kW")

                # Store in session state for slides
                st.session_state["ipals_data"] = {
                    "annual_gen_kwh": round(_total_gen),
                    "annual_demand_kwh": round(_total_demand),
                    "self_consumption_kwh": round(_total_self_consume),
                    "self_consumption_pct": round(_self_rate, 1) / 100,
                    "surplus_kwh": round(_total_surplus),
                    "co2_annual_t": round(_co2_t, 1),
                    "monthly_gen_kwh": [round(m) for m in _monthly_gen],
                    "hourly_rows": _hourly_rows,
                    "peak_demand_before_kw": round(_peak_before, 1),
                    "peak_demand_after_kw": round(_peak_after, 1),
                    "demand_cut_kw": round(_demand_cut, 1),
                }

                # 20-year projection
                with st.expander("20年間発電量推計", expanded=False):
                    _degradation = 0.005  # 0.5% annual degradation
                    _proj_rows = []
                    _total_20yr_gen = 0.0
                    _total_20yr_co2 = 0.0
                    for _yr in range(1, 21):
                        _decay = (1 - _degradation) ** (_yr - 1)
                        _yr_gen = _total_gen * _decay
                        _yr_co2 = _co2_t * _decay
                        _total_20yr_gen += _yr_gen
                        _total_20yr_co2 += _yr_co2
                        _proj_rows.append({
                            "年": _yr,
                            "発電量 (kWh)": f"{_yr_gen:,.0f}",
                            "自家消費 (kWh)": f"{_total_self_consume * _decay:,.0f}",
                            "余剰 (kWh)": f"{_total_surplus * _decay:,.0f}",
                            "CO2削減 (t)": f"{_yr_co2:.1f}",
                            "劣化率": f"{(1 - _decay) * 100:.1f}%",
                        })
                    import pandas as _pd
                    st.dataframe(_pd.DataFrame(_proj_rows), use_container_width=True, hide_index=True)

                    _tc1, _tc2 = st.columns(2)
                    with _tc1:
                        st.metric("20年間累計発電量", f"{_total_20yr_gen:,.0f} kWh")
                    with _tc2:
                        st.metric("20年間累計CO2削減", f"{_total_20yr_co2:,.1f} t-CO2")

                    # Store 20-year totals
                    st.session_state["ipals_data"]["total_20yr_gen_kwh"] = round(_total_20yr_gen)
                    st.session_state["ipals_data"]["total_20yr_co2_t"] = round(_total_20yr_co2, 1)
            else:
                st.error("CSVのパースに失敗しました。iPals出力形式を確認してください。")

    # ----- PPA Unit Price Auto-Calculation (PPA only) -----
    if not is_epc:
        with st.expander("💹 PPA単価自動試算（DSCR ≥ 1.30）", expanded=False):
            from proposal_generator.ppa_calc import (
                auto_calc_ppa,
                DEFAULT_MAINTENANCE_YEN_PER_KW,
                DEFAULT_INSURANCE_YEN_FIXED,
            )

            _ipals_now = st.session_state.get("ipals_data", {})
            _sc_y1 = _ipals_now.get("self_consumption_kwh", 0.0)
            _sur_y1 = _ipals_now.get("surplus_kwh", 0.0)

            _ac_col1, _ac_col2 = st.columns(2)
            with _ac_col1:
                _include_sur = st.checkbox(
                    "余剰売電収入を計上する（通常はRPRのためオフ）",
                    value=False,
                    key="ppa_include_surplus",
                )
                _target_dscr = st.number_input(
                    "目標DSCR",
                    min_value=1.0, max_value=3.0, value=1.30, step=0.05,
                    key="ppa_target_dscr",
                )
                with st.expander("O&M費用設定（Excelデフォルト値）", expanded=False):
                    _maint_per_kw = st.number_input(
                        "保守費 (円/kW/年)",
                        min_value=0, value=int(DEFAULT_MAINTENANCE_YEN_PER_KW), step=100,
                        key="ppa_maint_per_kw",
                        help=f"デフォルト: {DEFAULT_MAINTENANCE_YEN_PER_KW:,.0f} 円/kW/年",
                    )
                    _security_type = st.radio(
                        "保安管理業務委託費",
                        ["自社 (¥120,000/年)", "先方負担 (¥0)", "他社委託 (金額入力)"],
                        horizontal=True, key="ppa_security_type",
                    )
                    if "自社" in _security_type:
                        _insure_fixed = 120_000
                    elif "先方" in _security_type:
                        _insure_fixed = 0
                    else:
                        _insure_fixed = st.number_input(
                            "他社委託費 (円/年)",
                            min_value=0, value=120_000, step=10_000,
                            key="ppa_insure_fixed",
                        )
                    _om_annual = system_capacity * _maint_per_kw + _insure_fixed
                    st.caption(f"年間O&M合計: ¥{_om_annual:,.0f}")
            with _ac_col2:
                if _sc_y1 > 0:
                    st.metric("自家消費量（初年度）", f"{_sc_y1:,.0f} kWh")
                    st.metric("元本（販売価格 - 補助金）", f"¥{max(selling_price - subsidy_amount, 0):,.0f}")
                    st.metric("システム容量", f"{system_capacity:.1f} kW")
                else:
                    st.warning("iPalsデータをアップロードすると自動試算できます")

            if st.button("試算する", key="calc_ppa_btn", type="primary", disabled=(_sc_y1 <= 0 or selling_price <= 0)):
                _result = auto_calc_ppa(
                    self_consumption_y1_kwh=_sc_y1,
                    surplus_y1_kwh=_sur_y1,
                    selling_price=selling_price,
                    subsidy_amount=subsidy_amount,
                    lease_company=lease_company,
                    lease_rate_pct=lease_rate,
                    lease_years=int(lease_years),
                    contract_years=int(contract_years),
                    system_kw=system_capacity,
                    fit_price=surplus_price,
                    include_surplus=_include_sur,
                    target_dscr=_target_dscr,
                    maintenance_yen_per_kw=_maint_per_kw,
                    insurance_yen_fixed=_insure_fixed,
                )
                st.session_state["ppa_calc_result"] = _result

            _calc_res = st.session_state.get("ppa_calc_result")
            if _calc_res:
                _warns = _calc_res.get("warnings", [])
                for _w in _warns:
                    st.warning(_w)

                _is_loan_result = _calc_res.get("finance_type") == "loan"
                _r1, _r2, _r3, _r4, _r5 = st.columns(5)
                with _r1:
                    _pay_label = "年間元金返済額" if _is_loan_result else "リース年額"
                    st.metric(_pay_label, f"¥{_calc_res['annual_lease_payment']:,.0f}")
                with _r2:
                    st.metric("O&M年額", f"¥{_calc_res.get('annual_om_cost', 0):,.0f}")
                with _r3:
                    st.metric("実効金利", f"{_calc_res['effective_rate_pct']:.2f}%")
                with _r4:
                    st.metric("最小PPA単価（DSCR達成）", f"{_calc_res['min_ppa_price']:.1f} 円/kWh")
                with _r5:
                    _md = _calc_res.get("min_dscr")
                    st.metric("最小DSCR（最終年）", f"{_md:.3f}" if _md else "—")

                # IRR / NPV / Re-lease
                _irr_val = _calc_res.get("irr_pct")
                _npv_val = _calc_res.get("npv_yen")
                _release = _calc_res.get("re_lease_annual", 0)
                _fin_cols = st.columns(3)
                with _fin_cols[0]:
                    st.metric("IRR", f"{_irr_val:.2f}%" if _irr_val is not None else "—")
                with _fin_cols[1]:
                    st.metric("NPV", f"¥{_npv_val:,.0f}" if _npv_val is not None else "—")
                with _fin_cols[2]:
                    if _release > 0:
                        st.metric("再リース年額", f"¥{_release:,.0f}")
                    else:
                        st.metric("再リース", "—")

                # Bank loan additional cost breakdown
                if _is_loan_result:
                    _lr1, _lr2, _lr3, _lr4 = st.columns(4)
                    with _lr1:
                        st.metric("年間元金返済額", f"¥{_calc_res.get('annual_principal', 0):,.0f}")
                    with _lr2:
                        st.metric("初年度利息", f"¥{_calc_res.get('annual_interest_y1', 0):,.0f}")
                    with _lr3:
                        st.metric("火災保険料(年額)", f"¥{_calc_res.get('fire_insurance_annual', 0):,.0f}")
                    with _lr4:
                        st.metric("償却資産税(初年度)", f"¥{_calc_res.get('depreciation_tax_y1', 0):,.0f}")

                # Apply button
                if st.button(
                    f"この単価を適用 → {_calc_res['min_ppa_price']:.1f} 円/kWh",
                    key="apply_ppa_price",
                    type="secondary",
                ):
                    st.session_state["ppa_price_input"] = float(_calc_res["min_ppa_price"])
                    st.rerun()

                # Cashflow table
                _cf = _calc_res.get("cashflow_table", [])
                if _cf:
                    import pandas as pd
                    _df = pd.DataFrame(_cf)
                    _df.columns = ["年", "自家消費(kWh)", "余剰(kWh)", "PPA収入(円)",
                                   "余剰収入(円)", "収入合計(円)", "リース料(円)", "O&M(円)",
                                   "費用合計(円)", "純CF(円)", "DSCR"]
                    st.dataframe(
                        _df.style.format({
                            "自家消費(kWh)": "{:,.0f}",
                            "余剰(kWh)": "{:,.0f}",
                            "PPA収入(円)": "{:,.0f}",
                            "余剰収入(円)": "{:,.0f}",
                            "収入合計(円)": "{:,.0f}",
                            "リース料(円)": "{:,.0f}",
                            "O&M(円)": "{:,.0f}",
                            "費用合計(円)": "{:,.0f}",
                            "純CF(円)": "{:,.0f}",
                            "DSCR": lambda x: f"{x:.3f}" if x else "—",
                        }),
                        use_container_width=True,
                        height=300,
                    )

    # ----- FF (Fact Finding) -----
    with st.expander("📝 FF振り返り（前回ヒアリング結果）", expanded=False):
        st.caption("NEW_ff スライドに使用します")
        ff_col1, ff_col2 = st.columns(2)
        with ff_col1:
            ff_current = st.text_area(
                "現状・課題", height=90,
                placeholder="例：月間電気代80万円、ピーク時のデマンド問題あり",
            )
            ff_needs = st.text_area(
                "担当者のニーズ", height=80,
                placeholder="例：稟議を通すための根拠資料が必要",
            )
        with ff_col2:
            ff_mgmt = st.text_area(
                "経営者へのアピールポイント", height=80,
                placeholder="例：20年間の電気代削減額、ROI、環境対応",
            )
            ff_constraints = st.text_area(
                "制約・懸念事項", height=80,
                placeholder="例：屋根の耐荷重確認が必要、補助金期限2026/6",
            )

    # ----- Store all data in session state -----
    st.session_state["customer_data"] = {
        "proposal_type": "epc" if is_epc else "ppa",
        "company_name": company_name,
        "office_name": office_name,
        "address": address,
        "opp_id": st.session_state.get("sf_opp_id", ""),
        "snow_depth": snow_depth,
        "proposal_date": str(proposal_date),
        "company_size": company_size,
        "site_survey": site_survey,
        "tax_display": tax_display,
        # Panel types (list of dicts)
        "panels": panel_data_list,
        "panel_total_kw": total_panel_kw,
        "panel_total_count": total_panel_count,
        # PCS types (list of dicts)
        "pcs_list": pcs_data_list,
        "pcs_total_kw": total_pcs_kw,
        "pcs_total_count": total_pcs_count,
        # Battery types (list of dicts)
        "batteries": battery_data_list,
        "battery_total_kwh": total_battery_kwh,
        "battery_total_count": total_battery_count,
        # Pricing
        "kw_unit_cost": kw_unit_cost,
        "raw_cost": raw_cost,
        "gross_margin_pct": gross_margin_pct,
        "gross_profit": gross_profit,
        "selling_price": selling_price,
        "sales_commission_pct": sales_commission_pct,
        "sales_commission_amount": commission_amount,
        # Flat values for Excel compatibility
        "system_capacity_kw": system_capacity,
        "panel_watt": panel_data_list[0]["watt_per_unit"] if panel_data_list else 0,
        "panel_count": total_panel_count,
        "pcs_output_kw": total_pcs_kw,
        "battery_kwh": total_battery_kwh,
        "contract_years": int(contract_years),
        "ppa_unit_price": ppa_unit_price,
        "surplus_price": surplus_price,
        "demand_reduction_kw": demand_reduction,
        "subsidy_name": subsidy_name,
        "subsidy_amount": subsidy_amount,
        "lease_company": lease_company,
        "lease_rate": lease_rate,
        "lease_years": int(lease_years),
        # Current annual electricity cost (for PP7/PP8 savings calculation)
        "annual_cost": annual_elec_cost if annual_elec_cost > 0 else None,
        "elec_company": st.session_state.get("elec_company", ""),
        "elec_contract": st.session_state.get("elec_contract", ""),
        "contract_kw": st.session_state.get("contract_kw", 0),
        "annual_kwh": st.session_state.get("annual_kwh", 0),
        # Demand cut data
        "basic_rate_kw": st.session_state.get("_basic_rate_kw", 0),
        "power_factor_pct": st.session_state.get("power_factor_pct", 85),
        # CO2 emission factor
        "co2_emission_factor": st.session_state.get("_co2_emission_factor", 0.000453),
        # PPA calc results (if auto-calculated)
        "annual_lease_payment": st.session_state.get("ppa_calc_result", {}).get("annual_lease_payment", 0),
        "ppa_effective_rate_pct": st.session_state.get("ppa_calc_result", {}).get("effective_rate_pct", 0.0),
        "annual_om_cost": st.session_state.get("ppa_calc_result", {}).get("annual_om_cost", 0),
        "total_annual_cost": st.session_state.get("ppa_calc_result", {}).get("total_annual_cost", 0),
        "min_ppa_price": st.session_state.get("ppa_calc_result", {}).get("min_ppa_price", 0),
        "min_dscr": st.session_state.get("ppa_calc_result", {}).get("min_dscr", None),
        "cashflow_table": st.session_state.get("ppa_calc_result", {}).get("cashflow_table", []),
        "ppa_principal": st.session_state.get("ppa_calc_result", {}).get("principal", 0),
        "irr_pct": st.session_state.get("ppa_calc_result", {}).get("irr_pct"),
        "npv_yen": st.session_state.get("ppa_calc_result", {}).get("npv_yen"),
        "re_lease_annual": st.session_state.get("ppa_calc_result", {}).get("re_lease_annual", 0),
        # FF
        "ff_current_situation": ff_current,
        "ff_customer_needs": ff_needs,
        "ff_mgmt_needs": ff_mgmt,
        "ff_constraints": ff_constraints,
    }
    # Merge iPals data if available
    _ipals = st.session_state.get("ipals_data")
    if _ipals:
        st.session_state["customer_data"].update({
            "annual_gen_kwh": _ipals.get("annual_gen_kwh"),
            "self_consumption_kwh": _ipals.get("self_consumption_kwh"),
            "self_consumption_pct": _ipals.get("self_consumption_pct"),
            "surplus_kwh": _ipals.get("surplus_kwh"),
            "co2_annual_t": _ipals.get("co2_annual_t"),
            "monthly_gen_kwh": _ipals.get("monthly_gen_kwh"),
            "hourly_rows": _ipals.get("hourly_rows"),
            "peak_demand_before_kw": _ipals.get("peak_demand_before_kw"),
            "peak_demand_after_kw": _ipals.get("peak_demand_after_kw"),
            "demand_cut_kw": _ipals.get("demand_cut_kw"),
        })

    # Compute annual_saving for PP7/PP8/new_summary
    _cd = st.session_state["customer_data"]
    _annual_cost = _cd.get("annual_cost")
    _self_kwh = _cd.get("self_consumption_kwh")
    _ppa_price = _cd.get("ppa_unit_price", 0) or 0
    _annual_kwh = _cd.get("annual_kwh", 0) or 0
    _is_epc_calc = _cd.get("proposal_type") == "epc"

    if _annual_cost and _self_kwh:
        if _is_epc_calc:
            # EPC: annual saving = self-consumption × average electricity rate
            _avg_rate = float(_annual_cost) / float(_annual_kwh) if _annual_kwh > 0 else 0
            _annual_saving = float(_self_kwh) * _avg_rate
        elif float(_ppa_price) > 0:
            # PPA: annual saving = current cost for self-consumed portion - PPA cost
            _avg_rate = float(_annual_cost) / float(_annual_kwh) if _annual_kwh > 0 else 0
            _current_cost = float(_self_kwh) * _avg_rate
            _ppa_annual_cost = float(_self_kwh) * float(_ppa_price)
            _annual_saving = _current_cost - _ppa_annual_cost
        else:
            _annual_saving = 0

        if _annual_saving > 0:
            _cd["annual_saving"] = _annual_saving
            _cd["annual_cost_saving"] = _annual_saving
            # Simple payback period
            _sell_price = _cd.get("selling_price", 0)
            if _sell_price > 0:
                _cd["investment_recovery_yr"] = round(float(_sell_price) / _annual_saving, 1)

# =========================================================================
# Tab 3: Slide Composition
# =========================================================================

with tab3:
    st.subheader("スライド構成")

    # Filter profiles by proposal type selected in tab2
    _ptype = "epc" if is_epc else "ppa"
    filtered_profiles = {
        k: v for k, v in profiles.items()
        if v.get("proposal_type") == _ptype
    }

    col_left, col_right = st.columns([1, 2])

    with col_left:
        _fp_keys = list(filtered_profiles.keys())
        persona = st.radio(
            "ペルソナを選択してください",
            _fp_keys,
            key=f"persona_radio_{_ptype}",
            help="選択したペルソナに基づいて推奨スライド構成が自動設定されます",
        )
        # Fallback if persona not in filtered list (happens on PPA↔EPC switch)
        if persona not in filtered_profiles:
            persona = _fp_keys[0] if _fp_keys else None
        if persona:
            st.caption(filtered_profiles[persona]["description"])

    with col_right:
        if not persona:
            st.warning("プロファイルが見つかりません")
            default_slides = []
            optional_slides = []
        else:
            default_slides = filtered_profiles[persona]["slides"]
            optional_slides = filtered_profiles[persona].get("optional", [])

        st.write("**① 含めるスライドを選択**")
        checked_slides = []
        for sid in default_slides:
            info = catalog.get(sid, {})
            title = info.get("title", sid)
            if st.checkbox(f"{sid}  ─  {title}", value=True, key=f"chk_{sid}_{persona}"):
                checked_slides.append(sid)

        if optional_slides:
            with st.expander("オプション（追加可能）"):
                for sid in optional_slides:
                    info = catalog.get(sid, {})
                    title = info.get("title", sid)
                    if st.checkbox(
                        f"{sid}  ─  {title}", value=False, key=f"opt_{sid}_{persona}"
                    ):
                        checked_slides.append(sid)

        st.divider()
        st.write("**② ドラッグで順序を変更**")
        if checked_slides:
            # Include slide count + hash in key so list refreshes on checkbox changes
            _sort_key = f"sort_{persona}_{len(checked_slides)}_{hash(tuple(checked_slides))}"
            sorted_slides = sort_items(
                [
                    f"{sid}  ─  {catalog.get(sid, {}).get('title', sid)}"
                    for sid in checked_slides
                ],
                direction="vertical",
                custom_style=_SORTABLE_STYLE,
                key=_sort_key,
            )
            final_slides = [item.split("  ─  ")[0].strip() for item in sorted_slides]
            st.session_state["selected_slides"] = final_slides
        else:
            st.warning("スライドが選択されていません")
            st.session_state["selected_slides"] = []

# =========================================================================
# Tab 4: Generate
# =========================================================================

with tab4:
    st.subheader("PPTX生成")

    customer_data = st.session_state.get("customer_data", {})
    selected_slides = st.session_state.get("selected_slides", [])
    ipals_file_val = st.session_state.get("ipals_file")

    # Summary cards
    _is_epc_tab4 = customer_data.get("proposal_type") == "epc"
    sum_col1, sum_col2, sum_col3, sum_col4, sum_col5 = st.columns(5)
    with sum_col1:
        st.metric("取引先", customer_data.get("company_name") or "未選択")
    with sum_col2:
        st.metric("提案タイプ", "EPC（販売）" if _is_epc_tab4 else "PPA（第三者所有）")
    with sum_col3:
        st.metric("システム容量", f"{customer_data.get('system_capacity_kw', 0):,.1f} kW")
    with sum_col4:
        if _is_epc_tab4:
            sp = customer_data.get("selling_price", 0)
            cap = customer_data.get("system_capacity_kw", 0)
            kw_price = int(sp / cap) if sp and cap else 0
            st.metric("販売価格", f"¥{sp:,.0f}" if sp else "—",
                      delta=f"¥{kw_price:,}/kW" if kw_price else None, delta_color="off")
        else:
            st.metric("PPA単価", f"{customer_data.get('ppa_unit_price', 0):.1f} 円/kWh")
    with sum_col5:
        st.metric("スライド数", f"{len(selected_slides)} 枚")

    st.divider()

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.write("**顧客情報**")
        st.write(f"- 取引先名: {customer_data.get('company_name') or '（未選択）'}")
        st.write(f"- 商談: {customer_data.get('office_name') or '—'}")
        st.write(f"- 所在地: {customer_data.get('address') or '—'}")
        if not _is_epc_tab4:
            st.write(f"- 契約期間: {customer_data.get('contract_years', 20)} 年")

    with col_b:
        # Equipment summary
        st.write("**設備構成**")
        panels = customer_data.get("panels", [])
        for p in panels:
            if p["count"] > 0:
                st.write(
                    f"- パネル: {p.get('model') or '未指定'} "
                    f"{p['watt_per_unit']:.0f}W × {p['count']:,}枚 = {p['total_kw']:,.2f}kW"
                )
        pcs_list = customer_data.get("pcs_list", [])
        for q in pcs_list:
            if q["count"] > 0:
                st.write(
                    f"- PCS: {q.get('model') or '未指定'} "
                    f"{q['kw_per_unit']:.1f}kW × {q['count']:,}台 = {q['total_kw']:,.2f}kW"
                )
        batteries = customer_data.get("batteries", [])
        for b in batteries:
            if b["count"] > 0:
                st.write(
                    f"- 蓄電池: {b.get('model') or '未指定'} "
                    f"{b['kwh_per_unit']:.1f}kWh × {b['count']:,}台 = {b['total_kwh']:,.1f}kWh"
                )

        # Pricing summary
        rc = customer_data.get("raw_cost", 0)
        gp = customer_data.get("gross_profit", 0)
        sp = customer_data.get("selling_price", 0)
        if sp > 0:
            st.write("**価格**")
            st.write(f"- 原価: ¥{rc:,.0f}")
            st.write(f"- 粗利: ¥{gp:,.0f} ({customer_data.get('gross_margin_pct', 0):.1f}%)")
            st.write(f"- 販売価格: ¥{sp:,.0f}")
            ca = customer_data.get("sales_commission_amount", 0)
            if ca > 0:
                st.write(f"- 販売手数料: ¥{ca:,.0f}")

    with col_c:
        st.write("**スライド構成**")
        for i, sid in enumerate(selected_slides, 1):
            title = catalog.get(sid, {}).get("title", sid)
            st.write(f"{i}. {sid} ─ {title}")

    st.divider()

    use_excel = st.checkbox(
        "Excelで計算を実行する（xlwings / Windowsのみ）",
        value=True,
        help="チェックを外すと、フォーム入力値のみでPPTXを生成します",
    )

    generate_btn = st.button(
        "🚀  PPTX を生成",
        type="primary",
        disabled=not customer_data.get("company_name"),
    )

    if generate_btn:
        with st.spinner("生成中..."):
            data = dict(customer_data)

            if use_excel and EXCEL_PATH.exists():
                try:
                    from proposal_generator.excel_runner import (
                        CustomerInput,
                        run_excel_calculation,
                    )

                    ci = CustomerInput(
                        **{
                            k: v
                            for k, v in customer_data.items()
                            if k in CustomerInput.__dataclass_fields__
                        }
                    )
                    ipals_csv = None
                    if ipals_file_val:
                        ipals_csv = ipals_file_val.read().decode(
                            "utf-8-sig", errors="replace"
                        )

                    excel_out = run_excel_calculation(EXCEL_PATH, ci, ipals_csv=ipals_csv)
                    data.update({k: v for k, v in excel_out.items() if v is not None})
                    st.success("✅ Excel計算完了")
                except Exception as e:
                    st.warning(f"Excel計算をスキップしました: {e}")

            from proposal_generator.generator import generate_proposal

            with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as tmp:
                output_path = Path(tmp.name)

            try:
                generate_proposal(
                    slide_ids=selected_slides,
                    data=data,
                    output_path=output_path,
                )
                with open(output_path, "rb") as f:
                    pptx_bytes = f.read()

                company = customer_data.get("company_name", "提案") or "提案"
                _type_label = "EPC" if customer_data.get("proposal_type") == "epc" else "PPA"
                filename = (
                    f"{_type_label}提案_{company}_{customer_data.get('proposal_date', '')}.pptx"
                )

                st.success("✅ 生成完了！")
                st.session_state["last_pptx_bytes"] = pptx_bytes
                st.session_state["last_pptx_filename"] = filename
                st.download_button(
                    label="📥  PPTXをダウンロード",
                    data=pptx_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                )

                # Box upload
                from proposal_generator.box_client import is_available as _box_ok_t4
                _box_fid = st.session_state.get("box_proposal_folder_id")
                if _box_ok_t4() and _box_fid:
                    if st.button("📦 Boxにアップロード", key="box_upload_pptx"):
                        try:
                            from proposal_generator.box_client import upload_file as _box_up
                            _tmp_box = Path(tempfile.mktemp(suffix=".pptx"))
                            with open(_tmp_box, "wb") as _bf:
                                _bf.write(pptx_bytes)
                            _res = _box_up(_box_fid, _tmp_box, filename)
                            _tmp_box.unlink(missing_ok=True)
                            st.success(f"📦 Boxにアップロード完了: {_res['name']}")
                        except Exception as e:
                            st.error(f"Boxアップロードエラー: {e}")
                elif _box_ok_t4():
                    st.caption("📦 Tab 1でBoxフォルダを検索すると、ここからアップロードできます")

            except Exception as e:
                st.error(f"生成エラー: {e}")
                raise
            finally:
                output_path.unlink(missing_ok=True)
