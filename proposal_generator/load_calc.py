"""
load_calc.py - Load calculation parser for 積載荷重計算表 Excel.

Reads the 'まとめ' sheet from the load calculation Excel and returns
structured data for use in the layout slide (PP5).
"""
from __future__ import annotations

from typing import Any


def parse_load_calc_excel(file_or_path) -> dict[str, Any]:
    """Parse load calculation Excel and return structured data.

    Reads from the 'まとめ' sheet:
      Row 5: headers
      Row 6: summary values (B6-J6)
      Row 9+: component breakdown (panel model, frame info, wiring)

    Returns dict with keys:
    - panel_model: str (PV型番, e.g. "PV：Jinko585-J")
    - panel_count: int (パネル枚数)
    - panel_weight_kg: float (パネル重量)
    - frame_weight_kg: float (架台重量)
    - wiring_weight_kg: float (配線重量)
    - total_weight_kg: float (総重量)
    - panel_area_m2: float (パネル面積)
    - roof_area_m2: float (屋根面積)
    - load_per_panel_area: float (kg/m2 対パネル面積)
    - load_per_roof_area: float (kg/m2 対屋根面積)
    - panel_unit_weight_kg: float (PV単体重量)
    - frame_model: str (架台型番)
    - frame_count: int (架台本数)
    - frame_unit_weight_kg: float (架台単体重量)
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_or_path, data_only=True)

    # Find 'まとめ' sheet
    ws = None
    for name in wb.sheetnames:
        if "まとめ" in name:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(
            f"'まとめ' sheet not found. Available sheets: {wb.sheetnames}"
        )

    def _safe_float(val, default: float = 0.0) -> float:
        if val is None:
            return default
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    def _safe_int(val, default: int = 0) -> int:
        if val is None:
            return default
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return default

    def _safe_str(val, default: str = "") -> str:
        if val is None:
            return default
        return str(val).strip()

    # Row 6: summary values
    result = {
        "panel_count": _safe_int(ws["B6"].value),
        "panel_weight_kg": _safe_float(ws["C6"].value),
        "frame_weight_kg": _safe_float(ws["D6"].value),
        "wiring_weight_kg": _safe_float(ws["E6"].value),
        "total_weight_kg": _safe_float(ws["F6"].value),
        "panel_area_m2": _safe_float(ws["G6"].value),
        "roof_area_m2": _safe_float(ws["H6"].value),
        "load_per_panel_area": _safe_float(ws["I6"].value),
        "load_per_roof_area": _safe_float(ws["J6"].value),
    }

    # Row 9+: component breakdown
    # A9: PV model name
    result["panel_model"] = _safe_str(ws["A9"].value)
    # B9 or C9: panel count (some formats use B, some C)
    _pv_count = _safe_int(ws["B9"].value) or _safe_int(ws["C9"].value)
    if _pv_count > 0:
        result["panel_count"] = _pv_count  # override with detail if available

    # B10 or C10: panel unit weight
    result["panel_unit_weight_kg"] = (
        _safe_float(ws["B10"].value) or _safe_float(ws["C10"].value)
    )

    # A12+: frame info
    result["frame_model"] = _safe_str(ws["A12"].value)
    result["frame_count"] = _safe_int(ws["B12"].value) or _safe_int(ws["C12"].value)
    result["frame_unit_weight_kg"] = (
        _safe_float(ws["B13"].value) or _safe_float(ws["C13"].value)
    )

    wb.close()
    return result
